"""
一括請求書作成マクロ用ファイル生成スクリプト

生成されるファイル:
  BulkInvoiceTemplate.xlsx  ... マクロ管理ブック（設定・請求先リスト）
  請求書テンプレート.xlsx     ... 請求書のテンプレートファイル（コピー元）

既に独自のテンプレートファイルがある場合は 請求書テンプレート.xlsx の
生成をスキップし、設定シートのパスをそちらに変更してください。

使い方:
    pip install openpyxl
    python create_template.py
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ---- 共通スタイル ----

def thin_border(**overrides):
    sides = {k: Side(style="thin") for k in ("left", "right", "top", "bottom")}
    sides.update({k: Side(style=v) for k, v in overrides.items()})
    return Border(**sides)

HEADER_FILL  = PatternFill("solid", fgColor="1A3A6B")
SUBHEAD_FILL = PatternFill("solid", fgColor="D6E4F7")
AMOUNT_FILL  = PatternFill("solid", fgColor="F0F4FF")
TOTAL_FILL   = PatternFill("solid", fgColor="1A3A6B")
ODD_FILL     = PatternFill("solid", fgColor="F8F9FF")

WHITE   = Font(color="FFFFFF", bold=True)
BLUE    = Font(color="1A3A6B", bold=True)
GRAY    = Font(color="888888")
NUM_FMT = "#,##0"


# ================================================================
# 管理ブック: 設定シート
# ================================================================
def setup_setting_sheet(ws):
    ws.title = "設定"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 50

    ws["A1"] = "【設定】自社情報・マクロ設定"
    ws["A1"].font = Font(bold=True, size=13, color="1A3A6B")
    ws.merge_cells("A1:B1")

    rows = [
        ("A2",  "B2",  "会社名",                   "株式会社サンプル"),
        ("A3",  "B3",  "住所",                     "〒100-0001 東京都千代田区○○1-2-3"),
        ("A4",  "B4",  "TEL",                      "03-0000-0000"),
        ("A5",  "B5",  "振込先情報",                "○○銀行 △△支店 普通 1234567 カ）サンプル"),
        ("A6",  "B6",  "支払期限（日数）",           30),
        ("A7",  "B7",  "請求書番号プレフィックス",   "INV"),
        ("A8",  "B8",  "テンプレートファイルパス",   "請求書テンプレート.xlsx"),
        ("A9",  "B9",  "出力フォルダーパス",         "請求書出力"),
    ]

    for a_ref, b_ref, label, value in rows:
        lc = ws[a_ref]
        lc.value     = label
        lc.font      = Font(bold=True, color="555555")
        lc.fill      = SUBHEAD_FILL
        lc.border    = thin_border()
        lc.alignment = Alignment(vertical="center")

        vc = ws[b_ref]
        vc.value     = value
        vc.border    = thin_border()
        vc.alignment = Alignment(vertical="center", wrap_text=True)

    ws["B6"].number_format = "0"

    # 注記
    notes = [
        "A11", "※ 支払期限（日数）: 請求日から何日後を支払期限とするか（例: 30）",
        "A12", "※ テンプレートファイルパス: 相対パス可（BulkInvoiceTemplate.xlsx と同じフォルダー基準）",
        "A13", "※ 出力フォルダーパス: 相対パス可。存在しない場合は自動作成されます",
    ]
    for i in range(0, len(notes), 2):
        c = ws[notes[i]]
        c.value = notes[i + 1]
        c.font  = GRAY
        ws.merge_cells(f"{notes[i]}:B{notes[i][1:]}")


# ================================================================
# 管理ブック: 請求先リストシート
# ================================================================
def setup_list_sheet(ws):
    ws.title = "請求先リスト"

    ws["A1"] = "【請求先リスト】ここにデータを入力してください"
    ws["A1"].font = Font(bold=True, size=13, color="1A3A6B")
    ws.merge_cells("A1:U1")

    headers = [
        ("A", "No.",         6),
        ("B", "取引先名",    20),
        ("C", "請求日",      13),
        ("D", "支払期限",    13),
        ("E", "品目1",       20),
        ("F", "数量1",        8),
        ("G", "単価1",       12),
        ("H", "品目2",       18),
        ("I", "数量2",        8),
        ("J", "単価2",       12),
        ("K", "品目3",       18),
        ("L", "数量3",        8),
        ("M", "単価3",       12),
        ("N", "品目4",       18),
        ("O", "数量4",        8),
        ("P", "単価4",       12),
        ("Q", "品目5",       18),
        ("R", "数量5",        8),
        ("S", "単価5",       12),
        ("T", "消費税率(%)", 12),
        ("U", "備考",        25),
    ]

    for col_letter, label, width in headers:
        ws.column_dimensions[col_letter].width = width
        c = ws[f"{col_letter}2"]
        c.value     = label
        c.font      = WHITE
        c.fill      = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin_border()

    ws.row_dimensions[2].height = 30

    numeric_cols = {6, 7, 9, 10, 12, 13, 15, 16, 18, 19, 20}

    # サンプルデータ
    samples = [
        (1, "株式会社テスト商事",   "2026/1/31", "",           "システム開発費",    1,  500000, "保守費",   1, 50000, "", "", "", "", "", "", "", "", "", 10, ""),
        (2, "有限会社サンプル工業", "2026/1/31", "2026/2/28",  "製品A",            10,   15000, "製品B",    5, 20000, "送料", 1, 2000, "", "", "", "", "", "", 10, ""),
        (3, "○○株式会社",          "2026/2/28", "",           "コンサルティング費", 1,  200000, "",        "", "",     "",     "", "", "", "", "", "", "", "", 10, "月次レポート含む"),
    ]

    for row_idx, data in enumerate(samples, start=3):
        fill = ODD_FILL if row_idx % 2 == 1 else PatternFill()
        for col_idx, value in enumerate(data, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.border    = thin_border()
            c.fill      = fill
            if col_idx in numeric_cols:
                c.number_format = NUM_FMT
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(vertical="center")
        ws.cell(row=row_idx, column=3).number_format = "yyyy/m/d"
        ws.cell(row=row_idx, column=4).number_format = "yyyy/m/d"

    # 空入力行（10行）
    for row_idx in range(len(samples) + 3, len(samples) + 13):
        for col_idx in range(1, 22):
            c = ws.cell(row=row_idx, column=col_idx)
            c.border = thin_border()
            if col_idx in numeric_cols:
                c.number_format = NUM_FMT

    # AA列: 出力ファイルパス（マクロが記録）
    c = ws.cell(row=2, column=27, value="出力ファイルパス")
    c.font  = GRAY
    ws.column_dimensions["AA"].width = 45

    ws.freeze_panes = "B3"


# ================================================================
# テンプレートファイル（請求書テンプレート.xlsx）
# ================================================================
def create_invoice_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求書"

    # 列幅
    for col, w in [("A", 3), ("B", 26), ("C", 8), ("D", 4),
                   ("E", 10), ("F", 12), ("G", 14), ("H", 16)]:
        ws.column_dimensions[col].width = w

    # 行高
    for r in range(1, 28):
        ws.row_dimensions[r].height = 18
    ws.row_dimensions[6].height = 28
    ws.row_dimensions[9].height = 22

    # --- 自社情報（左上）---
    ws["B2"].font      = Font(bold=True, size=14, color="1A3A6B")
    ws["B3"].font      = GRAY
    ws["B4"].font      = GRAY

    # --- 請求書タイトル（右上）---
    ws["G1"] = "請　求　書"
    ws["G1"].font      = Font(bold=True, size=20, color="1A3A6B")
    ws["G1"].alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells("G1:H1")

    for ref, label in [("G2", "請求書番号："), ("G3", "発行日："), ("G4", "支払期限：")]:
        ws[ref].value     = label
        ws[ref].font      = Font(bold=True, color="555555")
        ws[ref].alignment = Alignment(horizontal="right", vertical="center")

    for ref in ("H2", "H3", "H4"):
        ws[ref].alignment = Alignment(horizontal="left", vertical="center")

    # 区切り線（行5）
    for col in range(1, 9):
        ws.cell(row=5, column=col).border = Border(
            bottom=Side(style="medium", color="1A3A6B"))

    # --- 取引先・合計金額（行6）---
    ws["B6"].font      = Font(bold=True, size=13)
    ws["B6"].alignment = Alignment(vertical="center")
    ws.merge_cells("B6:D6")

    ws["G6"] = "ご請求金額"
    ws["G6"].font      = Font(bold=True, color="555555")
    ws["G6"].fill      = SUBHEAD_FILL
    ws["G6"].alignment = Alignment(horizontal="center", vertical="center")
    ws["G6"].border    = thin_border()

    ws["H6"].font          = Font(bold=True, size=14, color="1A3A6B")
    ws["H6"].number_format = "¥#,##0－"
    ws["H6"].alignment     = Alignment(horizontal="right", vertical="center")
    ws["H6"].border        = thin_border()

    ws["B7"] = "下記のとおりご請求申し上げます。"
    ws["B7"].font      = Font(color="555555")
    ws.merge_cells("B7:H7")

    # 区切り線（行8）
    for col in range(1, 9):
        ws.cell(row=8, column=col).border = Border(
            bottom=Side(style="thin", color="AAAAAA"))

    # --- 品目ヘッダー（行9）---
    for ref, label in [("B9", "品　目"), ("E9", "数量"), ("F9", "単価"), ("G9", "金額（税抜）")]:
        c = ws[ref]
        c.value     = label
        c.font      = WHITE
        c.fill      = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()
    ws["H9"].fill   = HEADER_FILL
    ws["H9"].border = thin_border()
    ws.merge_cells("B9:D9")
    ws.merge_cells("G9:H9")

    # --- 品目行（行10〜14）---
    for row in range(10, 15):
        fill = ODD_FILL if row % 2 == 0 else PatternFill()
        for col in ("B", "C", "D"):
            ws[f"{col}{row}"].fill = fill
        ws.merge_cells(f"B{row}:D{row}")
        ws[f"B{row}"].border    = thin_border()
        ws[f"B{row}"].alignment = Alignment(vertical="center")
        ws[f"B{row}"].fill      = fill

        for col, fmt in [("E", NUM_FMT), ("F", NUM_FMT), ("G", NUM_FMT)]:
            c = ws[f"{col}{row}"]
            c.fill          = fill
            c.number_format = fmt
            c.alignment     = Alignment(horizontal="right", vertical="center")
            c.border        = thin_border()
        ws[f"H{row}"].fill   = fill
        ws[f"H{row}"].border = thin_border()
        ws.merge_cells(f"G{row}:H{row}")

    # 区切り線（行15）
    for col in range(1, 9):
        ws.cell(row=15, column=col).border = Border(
            bottom=Side(style="medium", color="1A3A6B"),
            top=Side(style="thin"))

    # --- 集計エリア（行16〜18）---
    for row, label, val_ref, is_total in [
        (16, "小　計",      "G16", False),
        (17, "消費税（10%）", "G17", False),
        (18, "合計（税込）",  "G18", True),
    ]:
        ws.row_dimensions[row].height = 22

        lc = ws[f"F{row}"]
        lc.value     = label
        lc.font      = WHITE if is_total else Font(bold=True, color="555555")
        lc.fill      = TOTAL_FILL if is_total else SUBHEAD_FILL
        lc.alignment = Alignment(horizontal="right", vertical="center")
        lc.border    = thin_border()

        vc = ws[val_ref]
        vc.font          = WHITE if is_total else Font(bold=True)
        vc.fill          = TOTAL_FILL if is_total else AMOUNT_FILL
        vc.number_format = NUM_FMT
        vc.alignment     = Alignment(horizontal="right", vertical="center")
        vc.border        = thin_border()
        ws[f"H{row}"].fill   = TOTAL_FILL if is_total else AMOUNT_FILL
        ws[f"H{row}"].border = thin_border()
        ws.merge_cells(f"G{row}:H{row}")

    # --- 振込先（行20〜21）---
    ws["B20"] = "【振込先】"
    ws["B20"].font = Font(bold=True, color="1A3A6B")
    ws.merge_cells("B20:H20")
    ws.merge_cells("B21:H21")
    ws["B21"].alignment = Alignment(vertical="center", wrap_text=True)

    ws.row_dimensions[22].height = 6
    for col in range(2, 9):
        ws.cell(row=22, column=col).border = Border(
            bottom=Side(style="thin", color="CCCCCC"))

    # --- 備考（行23〜24）---
    ws["B23"] = "【備考】"
    ws["B23"].font = Font(bold=True, color="1A3A6B")
    ws.merge_cells("B23:H23")
    ws.merge_cells("B24:H24")
    ws.row_dimensions[24].height = 45
    ws["B24"].alignment = Alignment(wrap_text=True, vertical="top")

    # 外枠
    for row in range(1, 25):
        left_cell  = ws.cell(row=row, column=1)
        right_cell = ws.cell(row=row, column=8)
        left_cell.border = Border(
            left=Side(style="medium"),
            right=left_cell.border.right,
            top=left_cell.border.top,
            bottom=left_cell.border.bottom,
        )
        right_cell.border = Border(
            left=right_cell.border.left,
            right=Side(style="medium"),
            top=right_cell.border.top,
            bottom=right_cell.border.bottom,
        )
    # 最上段・最下段
    for col in range(1, 9):
        c = ws.cell(row=1, column=col)
        c.border = Border(
            left=c.border.left, right=c.border.right,
            top=Side(style="medium"), bottom=c.border.bottom)
        c = ws.cell(row=24, column=col)
        c.border = Border(
            left=c.border.left, right=c.border.right,
            top=c.border.top, bottom=Side(style="medium"))

    # 印刷設定
    ws.page_setup.orientation  = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize    = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage    = True
    ws.page_setup.fitToWidth   = 1
    ws.page_setup.fitToHeight  = 1
    ws.print_area               = "A1:H26"
    ws.sheet_view.showGridLines = False

    output = "請求書テンプレート.xlsx"
    wb.save(output)
    return output


# ================================================================
# メイン
# ================================================================
def main():
    # 1. 管理ブック（BulkInvoiceTemplate.xlsx）を生成
    wb = openpyxl.Workbook()
    ws_setting = wb.active
    ws_list    = wb.create_sheet("請求先リスト")

    setup_setting_sheet(ws_setting)
    setup_list_sheet(ws_list)

    wb.active = ws_setting
    wb.save("BulkInvoiceTemplate.xlsx")
    print("✅ BulkInvoiceTemplate.xlsx を生成しました（設定・請求先リスト）")

    # 2. 請求書テンプレートファイルを生成
    tpl = create_invoice_template()
    print(f"✅ {tpl} を生成しました（請求書のコピー元テンプレート）")

    print()
    print("【次の手順】")
    print("  1. BulkInvoiceTemplate.xlsx を Excel で開く")
    print("  2. Alt+F11 → ファイル → ファイルのインポート → BulkInvoiceMacro.bas")
    print("  3. 「設定」シートを確認（独自テンプレートがある場合は B8 のパスを変更）")
    print("  4. 「請求先リスト」シートに請求データを入力")
    print("  5. Alt+F8 → 一括請求書作成 → 実行")
    print()
    print("  ※ 独自テンプレートを使う場合:")
    print("     設定シート B8 に独自テンプレートのパスを入力してください")
    print("     （例: C:\\Users\\user\\Documents\\my_invoice_template.xlsx）")
    print("     VBA 定数 TPL_* を独自テンプレートのセル位置に合わせて変更してください")


if __name__ == "__main__":
    main()
